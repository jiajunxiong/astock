import csv
import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook("zz500_1.xlsx")
sheets = wb.get_sheet_names()

for sheet in sheets:
    print (sheet)
    """
    ws = wb.get_sheet_by_name(sheet)
    name = ws['A1'].value
    print (name)
    ws['A2'].value = "DATE"
    ws['A3'].value = datetime(2018, 1, 2)
    
    path = "./zz500/"
    with open(path+name[:-3]+'.csv', 'w', newline="") as f:  
        counter_row = 0
        c = csv.writer(f)
        for r in ws.rows:
            if counter_row != 0:
                cell_list = [cell.value for cell in r]
                if counter_row == 1:    
                    c.writerow(cell_list)
                else:
                    cell_list[0] = cell_list[0].strftime("%Y-%m-%d")
                    c.writerow(cell_list)
            counter_row += 1
    """